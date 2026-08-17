import io
from datetime import date, datetime
from typing import Optional

from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.waste import Waste
from app.models.textile_analysis import ImageAnalysis, MaterialClassification, WasteClassification, RecyclingRecommendation
from app.models.sustainability_models import SustainabilityMetrics, EnvironmentalImpact, CircularityAnalysis, WasteScore

BRAND_TEAL = colors.HexColor("#1F6F5C")
BRAND_INK = colors.HexColor("#16211C")
BRAND_MUTED = colors.HexColor("#5C6B62")


def _apply_date_range(query, column, start_date: Optional[date], end_date: Optional[date]):
    if start_date:
        query = query.filter(column >= start_date)
    if end_date:
        query = query.filter(column <= end_date)
    return query


def _fetch_report_data(db: Session, report_type: str, start_date, end_date, user):
    """Returns (title, headers, rows, kpi_summary) for the requested report type."""
    if report_type == "waste_classification":
        q = db.query(WasteClassification)
        q = _apply_date_range(q, WasteClassification.created_at, start_date, end_date)
        rows_db = q.order_by(WasteClassification.created_at.desc()).limit(500).all()
        headers = ["Date", "Filename", "Category", "Condition", "Contamination %", "Recyclability %"]
        rows = []
        for r in rows_db:
            img = db.query(ImageAnalysis).filter(ImageAnalysis.id == r.image_id).first()
            rows.append([
                str(r.created_at.date()), img.filename if img else "—", r.waste_category,
                r.waste_condition or "—", f"{r.contamination_percentage:.0f}%", f"{r.recyclability_percentage:.0f}%",
            ])
        kpi = {"Total Records": len(rows), "Avg Recyclability": f"{(sum(r.recyclability_percentage or 0 for r in rows_db) / len(rows_db)):.1f}%" if rows_db else "—"}
        return "Waste Classification Report", headers, rows, kpi

    if report_type == "recycling":
        q = db.query(RecyclingRecommendation)
        q = _apply_date_range(q, RecyclingRecommendation.created_at, start_date, end_date)
        rows_db = q.order_by(RecyclingRecommendation.created_at.desc()).limit(500).all()
        headers = ["Date", "Best Method", "Sustainability Score", "Environmental Impact Score"]
        rows = [[str(r.created_at.date()), r.best_recycling_method, f"{r.sustainability_score:.0f}", f"{r.environmental_impact_score:.0f}"] for r in rows_db]
        kpi = {"Total Recommendations": len(rows)}
        return "Recycling Report", headers, rows, kpi

    if report_type == "sustainability":
        q = db.query(SustainabilityMetrics)
        q = _apply_date_range(q, SustainabilityMetrics.created_at, start_date, end_date)
        rows_db = q.order_by(SustainabilityMetrics.created_at.desc()).limit(500).all()
        headers = ["Date", "Material", "Weight (kg)", "Sustainability Index", "Rating", "Diverted %"]
        rows = [[str(r.created_at.date()), r.material_type, r.weight_kg, f"{r.sustainability_index:.0f}", r.sustainability_rating, f"{r.diverted_percentage:.0f}%"] for r in rows_db]
        kpi = {"Total Batches": len(rows), "Avg Index": f"{(sum(r.sustainability_index or 0 for r in rows_db) / len(rows_db)):.1f}" if rows_db else "—"}
        return "Sustainability Report", headers, rows, kpi

    if report_type == "environmental_impact":
        q = db.query(EnvironmentalImpact)
        q = _apply_date_range(q, EnvironmentalImpact.created_at, start_date, end_date)
        rows_db = q.order_by(EnvironmentalImpact.created_at.desc()).limit(500).all()
        headers = ["Date", "CO2 Saved (kg)", "Water Saved (L)", "Landfill Saved (kg)", "Rating"]
        rows = [[str(r.created_at.date()), r.co2_saved, r.water_saved, r.landfill_saved, r.rating] for r in rows_db]
        kpi = {
            "Total CO2 Saved": f"{sum(r.co2_saved or 0 for r in rows_db):.0f} kg",
            "Total Water Saved": f"{sum(r.water_saved or 0 for r in rows_db):,.0f} L",
        }
        return "Environmental Impact Report", headers, rows, kpi

    if report_type == "circular_economy":
        q = db.query(CircularityAnalysis)
        q = _apply_date_range(q, CircularityAnalysis.created_at, start_date, end_date)
        rows_db = q.order_by(CircularityAnalysis.created_at.desc()).limit(500).all()
        headers = ["Date", "Score", "Utilization %", "Optimization %", "Category"]
        rows = [[str(r.created_at.date()), f"{r.score:.0f}", f"{r.utilization:.0f}%", f"{r.optimization:.0f}%", r.category] for r in rows_db]
        kpi = {"Total Analyses": len(rows)}
        return "Circular Economy Report", headers, rows, kpi

    raise ValueError(f"Unknown report type: {report_type}")


def generate_pdf_report(db: Session, report_type: str, start_date, end_date, user) -> bytes:
    title, headers, rows, kpi = _fetch_report_data(db, report_type, start_date, end_date, user)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40, leftMargin=40, rightMargin=40)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("TIHeading", parent=styles["Heading1"], textColor=BRAND_TEAL, fontSize=18, spaceAfter=4)
    meta_style = ParagraphStyle("TIMeta", parent=styles["Normal"], textColor=BRAND_MUTED, fontSize=9)

    elements = []
    elements.append(Paragraph("TextileIntel", ParagraphStyle("Brand", parent=styles["Heading2"], textColor=BRAND_INK, fontSize=14)))
    elements.append(Paragraph(title, title_style))
    date_range_str = f"{start_date or 'All time'} to {end_date or 'present'}"
    elements.append(Paragraph(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC · Range: {date_range_str} · By: {user.email}", meta_style))
    elements.append(Spacer(1, 16))

    if kpi:
        kpi_data = [list(kpi.keys()), list(kpi.values())]
        kpi_table = Table(kpi_data, hAlign="LEFT")
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_TEAL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E3E1D8")),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 20))

    if rows:
        table_data = [headers] + rows
        data_table = Table(table_data, repeatRows=1)
        data_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E3E1D8")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F3")]),
        ]))
        elements.append(data_table)
    else:
        elements.append(Paragraph("No records found for the selected range.", styles["Normal"]))

    doc.build(elements)
    return buffer.getvalue()


def generate_excel_report(db: Session, report_type: str, start_date, end_date, user) -> bytes:
    title, headers, rows, kpi = _fetch_report_data(db, report_type, start_date, end_date, user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_fill = PatternFill(start_color="1F6F5C", end_color="1F6F5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14, color="16211C")
    ws.append([f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"])
    ws.append([])

    if kpi:
        for k, v in kpi.items():
            ws.append([k, v])
        ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append(headers)
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()